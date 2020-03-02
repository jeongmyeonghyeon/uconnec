"""
[아크릴] 실내 이미지 바운딩 (1636) 주간보고
"""
from crowdworks.commonUtil import *

import os
import shutil
import json
import uuid
import openpyxl


def switch(x):
    return {
        2035: "발코니_베란다/",
        2036: "욕실/",
        2037: "침실/",
        2038: "옷방/",
        2039: "식사실/",
        2040: "현관/",
        2041: "아이방/",
        2042: "주방/",
        2043: "거실/",
        2044: "서재_공부방/",
        2045: "다용도실/"
    }.get(x, False)


item_list = {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "무선충전기": 0, "티비(모니터포함)": 0, "사운드바": 0,
             "스피커": 0, "프린터": 0, "노트북": 0, "마우스": 0, "키보드": 0, "냉장고": 0, "오븐": 0, "전자레인지": 0, "인덕션": 0, "식기세척기": 0,
             "스탠드에어컨": 0, "벽걸이에어컨": 0, "시스템에어컨": 0, "드럼세탁기": 0, "전자동세탁기": 0, "건조기": 0, "의류관리기": 0, "무선 청소기": 0,
             "유선 청소기": 0, "로봇청소기": 0, "공기청정기": 0, "가습기": 0, "헤어드라이어": 0, "침대": 0, "화장대": 0, "스탠드옷걸이": 0, "옷장": 0,
             "서랍": 0, "놀이매트": 0, "책상": 0, "의자": 0, "책장": 0, "소파": 0, "TV 장식장": 0, "리모콘": 0, "카펫": 0, "피아노": 0,
             "실내자전기": 0, "선풍기": 0, "테이블": 0, "싱크볼": 0, "커피머신": 0, "조리기구": 0, "토스터기": 0, "밥솥": 0, "욕조": 0, "세면대": 0,
             "샤워기": 0, "변기": 0, "사람": 0, "강아지": 0, "고양이": 0}

task_id_list = "2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044, 2045"

sql_result_json = "SELECT result_json FROM TB_PRJ_DATA WHERE prj_idx in (" + task_id_list + ") AND prog_state_cd='ALL_FINISHED'"
print(sql_result_json)

result_json = getDatabaseData(sql_result_json, "prd", "mhjeong", "cworks@34")

for result in result_json:
    # str > dict
    result_dict = json.loads(result[0])

    print('result_dict: ', result_dict['data'])

    # for i in result_dict['data']:
    #     print('i: ', i['data'][0]['value'])

    # for label in result_dict['label-list']:
    #     try:
    #         for item in item_list:
    #             if label['label'] == item:
    #                 item_list[item] = item_list[item] + 1
    #     except KeyError:
    #         print(result_dict['fileId'])
    #         # prj_idx 16876106 의 result_json['label-list'] 에 'label' key 가 없당...
    #         pass
    for label in result_dict['data']:
        try:
            for item in item_list:
                if label['data'][0]['value'] == item:
                    item_list[item] = item_list[item] + 1
        except KeyError:
            print(result_dict['data'][0]['id'])
            # prj_idx 16876106 의 result_json['label-list'] 에 'label' key 가 없당...
            pass

# workbook 생성
wb = openpyxl.Workbook()
wb_detail = wb.create_sheet('상세')
index = 0
for item in item_list:
    index = index + 1

    name_key = "A" + str(index)
    value_key = "B" + str(index)

    wb['Sheet'][name_key] = item
    wb['Sheet'][value_key] = item_list[item]

# wb.close()
# wb.save('아크릴_데이터_수량_요약.xlsx')

# 상세는 아예 따로 돌릴깡...
item_list_detail = {
    "2035": {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "드럼세탁기": 0, "전자동세탁기": 0, "건조기": 0,
             "무선 청소기": 0, "유선 청소기": 0, "로봇청소기": 0, "서랍": 0, "리모콘": 0, "카펫": 0, "실내자전기": 0, "선풍기": 0, "사람": 0, "강아지": 0,
             "고양이": 0},
    "2036": {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "스피커": 0, "드럼세탁기": 0, "전자동세탁기": 0,
             "건조기": 0, "헤어드라이어": 0, "서랍": 0, "리모콘": 0, "욕조": 0, "세면대": 0, "샤워기": 0, "변기": 0, "사람": 0, "강아지": 0,
             "고양이": 0},
    "2037": {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "무선충전기": 0, "티비(모니터포함)": 0, "사운드바": 0,
             "스피커": 0, "프린터": 0, "노트북": 0, "마우스": 0, "키보드": 0, "스탠드에어컨": 0, "벽걸이에어컨": 0, "시스템에어컨": 0, "건조기": 0,
             "의류관리기": 0, "무선 청소기": 0, "유선 청소기": 0, "로봇청소기": 0, "공기청정기": 0, "가습기": 0, "헤어드라이어": 0, "침대": 0, "화장대": 0,
             "스탠드옷걸이": 0, "옷장": 0, "서랍": 0, "놀이배트": 0, "책상": 0, "의자": 0, "책장": 0, "소파": 0, "TV 장식장": 0, "리모콘": 0,
             "카펫": 0, "피아노": 0, "실내자전기": 0, "선풍기": 0, "커피머신": 0, "사람": 0, "강아지": 0, "고양이": 0},
    "2038": {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "무선충전기": 0, "스피커": 0, "노트북": 0, "마우스": 0,
             "키보드": 0, "스탠드에어컨": 0, "벽걸이에어컨": 0, "시스템에어컨": 0, "건조기": 0, "의류관리기": 0, "무선 청소기": 0, "유선 청소기": 0,
             "로봇청소기": 0, "공기청정기": 0, "가습기": 0, "헤어드라이어": 0, "화장대": 0, "스탠드옷걸이": 0, "옷장": 0, "서랍": 0, "책상": 0, "의자": 0,
             "책장": 0, "소파": 0, "리모콘": 0, "카펫": 0, "피아노": 0, "실내자전기": 0, "선풍기": 0, "사람": 0, "강아지": 0, "고양이": 0, },
    "2039": {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "무선충전기": 0, "스피커": 0, "냉장고": 0, "오븐": 0,
             "전자레인지": 0, "무선 청소기": 0, "유선 청소기": 0, "로봇청소기": 0, "공기청정기": 0, "가습기": 0, "서랍": 0, "리모콘": 0, "카펫": 0,
             "선풍기": 0, "테이블": 0, "싱크볼": 0, "커피머신": 0, "조리기구": 0, "토스터기": 0, "밥솥": 0, "사람": 0, "강아지": 0, "고양이": 0},
    "2040": {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "무선충전기": 0, "티비(모니터포함)": 0, "사운드바": 0,
             "스피커": 0, "프린터": 0, "노트북": 0, "마우스": 0, "키보드": 0, "냉장고": 0, "오븐": 0, "전자레인지": 0, "식기세척기": 0, "스탠드에어컨": 0,
             "벽걸이에어컨": 0, "시스템에어컨": 0, "드럼세탁기": 0, "전자동세탁기": 0, "건조기": 0, "의류관리기": 0, "무선 청소기": 0, "유선 청소기": 0,
             "로봇청소기": 0, "공기청정기": 0, "가습기": 0, "헤어드라이어": 0, "화장대": 0, "스탠드옷걸이": 0, "옷장": 0, "서랍": 0, "놀이배트": 0, "책상": 0,
             "의자": 0, "책장": 0, "소파": 0, "TV 장식장": 0, "리모콘": 0, "카펫": 0, "피아노": 0, "실내자전기": 0, "선풍기": 0, "테이블": 0,
             "커피머신": 0, "조리기구": 0, "토스터기": 0, "밥솥": 0, "사람": 0, "강아지": 0, "고양이": 0},
    "2041": {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "무선충전기": 0, "티비(모니터포함)": 0, "사운드바": 0,
             "스피커": 0, "스탠드에어컨": 0, "벽걸이에어컨": 0, "시스템에어컨": 0, "건조기": 0, "의류관리기": 0, "무선 청소기": 0, "유선 청소기": 0,
             "로봇청소기": 0, "공기청정기": 0, "가습기": 0, "헤어드라이어": 0, "스탠드옷걸이": 0, "옷장": 0, "서랍": 0, "놀이배트": 0, "소파": 0, "리모콘": 0,
             "피아노": 0, "실내자전기": 0, "선풍기": 0, "사람": 0, "강아지": 0, "고양이": 0},
    "2042": {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "무선충전기": 0, "티비(모니터포함)": 0, "스피커": 0,
             "냉장고": 0, "오븐": 0, "전자레인지": 0, "인덕션": 0, "식기세척기": 0, "시스템에어컨": 0, "무선 청소기": 0, "유선 청소기": 0, "로봇청소기": 0,
             "공기청정기": 0, "가습기": 0, "서랍": 0, "리모콘": 0, "카펫": 0, "피아노": 0, "실내자전기": 0, "선풍기": 0, "싱크볼": 0, "커피머신": 0,
             "조리기구": 0, "토스터기": 0, "밥솥": 0, "사람": 0, "강아지": 0, "고양이": 0},
    "2043": {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "무선충전기": 0, "티비(모니터포함)": 0, "사운드바": 0,
             "스피커": 0, "프린터": 0, "노트북": 0, "마우스": 0, "키보드": 0, "냉장고": 0, "오븐": 0, "전자레인지": 0, "스탠드에어컨": 0, "벽걸이에어컨": 0,
             "시스템에어컨": 0, "건조기": 0, "의류관리기": 0, "무선 청소기": 0, "유선 청소기": 0, "로봇청소기": 0, "공기청정기": 0, "가습기": 0, "헤어드라이어": 0,
             "스탠드옷걸이": 0, "서랍": 0, "놀이배트": 0, "책상": 0, "의자": 0, "책장": 0, "소파": 0, "TV 장식장": 0, "리모콘": 0, "카펫": 0,
             "피아노": 0, "실내자전기": 0, "선풍기": 0, "커피머신": 0, "토스터기": 0, "사람": 0, "강아지": 0, "고양이": 0},
    "2044": {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "무선충전기": 0, "티비(모니터포함)": 0, "사운드바": 0,
             "스피커": 0, "프린터": 0, "노트북": 0, "마우스": 0, "키보드": 0, "스탠드에어컨": 0, "벽걸이에어컨": 0, "시스템에어컨": 0, "무선 청소기": 0,
             "유선 청소기": 0, "로봇청소기": 0, "공기청정기": 0, "가습기": 0, "헤어드라이어": 0, "스탠드옷걸이": 0, "옷장": 0, "서랍": 0, "놀이배트": 0,
             "책상": 0, "의자": 0, "책장": 0, "소파": 0, "리모콘": 0, "카펫": 0, "피아노": 0, "실내자전기": 0, "선풍기": 0, "커피머신": 0, "사람": 0,
             "강아지": 0, "고양이": 0},
    "2045": {"스마트폰": 0, "태블릿": 0, "스마트워치": 0, "무선이어폰": 0, "헤드폰": 0, "VR기기": 0, "무선충전기": 0, "스피커": 0, "냉장고": 0, "오븐": 0,
             "전자레인지": 0, "식기세척기": 0, "드럼세탁기": 0, "전자동세탁기": 0, "건조기": 0, "의류관리기": 0, "무선 청소기": 0, "유선 청소기": 0,
             "로봇청소기": 0, "헤어드라이어": 0, "스탠드옷걸이": 0, "서랍": 0, "리모콘": 0, "카펫": 0, "커피머신": 0, "조리기구": 0, "토스터기": 0, "밥솥": 0,
             "사람": 0, "강아지": 0, "고양이": 0}
}
key_first_letter = {
    "2035": ["A", "B"],
    "2036": ["C", "D"],
    "2037": ["E", "F"],
    "2038": ["G", "H"],
    "2039": ["I", "J"],
    "2040": ["K", "L"],
    "2041": ["M", "N"],
    "2042": ["O", "P"],
    "2043": ["Q", "R"],
    "2044": ["S", "T"],
    "2045": ["U", "V"]
}
task_id_list_detail = [2035, 2036, 2037, 2038, 2039, 2040, 2041, 2042, 2043, 2044, 2045]
# task_id_list_detail = [1933]

for task_id in task_id_list_detail:
    t_id = str(task_id)
    sql_result_json_detail = "SELECT result_json FROM TB_PRJ_DATA WHERE prj_idx =" + t_id + " AND prog_state_cd='ALL_FINISHED'"
    result_json_detail = getDatabaseData(sql_result_json_detail, "prd", "mhjeong", "cworks@34")

    for result_detail in result_json_detail:
        result_detail_dict = json.loads(result_detail[0])

        for label in result_detail_dict['data']:
            try:
                for item_detail in item_list_detail[t_id]:
                    if label['data'][0]['value'] == item_detail:
                        item_list_detail[t_id][item_detail] = item_list_detail[t_id][item_detail] + 1
            except KeyError:
                print("keyError", result_detail_dict['data'][0]['id'])
                pass

    # for label in result_dict['data']:
    #     try:
    #         for item in item_list:
    #             if label['data'][0]['value'] == item:
    #                 item_list[item] = item_list[item] + 1
    #     except KeyError:
    #         print(result_dict['data'][0]['id'])
    #         # prj_idx 16876106 의 result_json['label-list'] 에 'label' key 가 없당...
    #         pass

    # wb = openpyxl.Workbook('아크릴_데이터_수량_요약.xlsx')
    index = 1

    wb_detail[key_first_letter[t_id][0] + str(index)] = switch(task_id)

    for item_detail in item_list_detail[t_id]:
        index = index + 1

        name_key_detail = key_first_letter[t_id][0] + str(index)
        value_key_detail = key_first_letter[t_id][1] + str(index)

        wb_detail[name_key_detail] = item_detail
        wb_detail[value_key_detail] = item_list_detail[t_id][item_detail]

    wb.close()
    wb.save('아크릴_데이터_수량_요약_BOUNDING.xlsx')
